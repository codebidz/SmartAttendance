import cv2
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, Response
import os
import mimetypes
from deepface import DeepFace
import numpy as np

# ============================
# Flask Setup
# ============================
mimetypes.init(files=[])
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"))

# ============================
# Camera Setup
# ============================
def get_camera_source():
    cap = cv2.VideoCapture(0)
    if cap.isOpened():
        print("Using laptop webcam")
        return cap
    cap.release()
    raise Exception("No camera source available")

camera_cap = get_camera_source()

# ============================
# Dataset Embeddings
# ============================
#DATASET_PATH = "dataset"
DATASET_PATH = os.path.join(BASE_DIR, "DataSet")
known_embeddings = []
known_face_names = []
known_face_rollnos = []

if os.path.exists(DATASET_PATH):
    for student_name in os.listdir(DATASET_PATH):
        student_folder = os.path.join(DATASET_PATH, student_name)
        if os.path.isdir(student_folder):
            for rollno in os.listdir(student_folder):
                roll_folder = os.path.join(student_folder, rollno)
                if os.path.isdir(roll_folder):
                    for img in os.listdir(roll_folder):
                        path = os.path.join(roll_folder, img)
                        if img.lower().endswith((".jpg", ".jpeg", ".png")):
                            try:
                                emb = DeepFace.represent(
                                    img_path=path,
                                    model_name="ArcFace",
                                    detector_backend="retinaface",
                                    enforce_detection=True
                                )[0]["embedding"]
                                known_embeddings.append(np.array(emb))
                                known_face_names.append(student_name)
                                known_face_rollnos.append(rollno)
                                print(f"Loaded embedding: {student_name} ({rollno})")
                            except Exception as e:
                                print(f"[ERROR] {e}")
else:
    print("Dataset folder not found")

# ============================
# Database + Excel Setup
# ============================
db_path = os.path.join(BASE_DIR, "attendance.db")
conn = sqlite3.connect(db_path, check_same_thread=False)
cursor = conn.cursor()
cursor.execute("DROP TABLE IF EXISTS attendance")
cursor.execute("""
CREATE TABLE attendance (
    name TEXT,
    rollno TEXT,
    date TEXT,
    time TEXT
)
""")
conn.commit()

excel_file = os.path.join(BASE_DIR, "attendance.xlsx")
if not os.path.exists(excel_file):
    wb = Workbook()
    wb.remove(wb.active)
    ws_present = wb.create_sheet("Present")
    ws_present.append(["Name", "Roll No", "Date", "Time", "Status"])
    ws_absent = wb.create_sheet("Absent")
    ws_absent.append(["Roll No", "Name"])
    wb.save(excel_file)

attendance_marked = set()

def mark_attendance(name, rollno):
    key = f"{name}_{rollno}"
    if key in attendance_marked:
        return
    now = datetime.now()
    date_string = now.strftime("%d-%m-%Y")
    time_string = now.strftime("%I:%M:%S %p")
    cursor.execute("INSERT INTO attendance VALUES (?, ?, ?, ?)",
                   (name, rollno, date_string, time_string))
    conn.commit()
    wb = load_workbook(excel_file)
    ws_present = wb["Present"]
    ws_present.append([name, rollno, date_string, time_string, "Present"])
    wb.save(excel_file)
    attendance_marked.add(key)
    print(f"Attendance marked: {name} ({rollno})")

# ============================
# Camera Streaming with RetinaFace
# ============================
'''def gen_frames():
    while True:
        success, frame = camera_cap.read()
        if not success:
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        try:
            # Detect and crop face
            detections = DeepFace.detectFace(
                rgb_frame,
                detector_backend="retinaface",
                enforce_detection=False
            )
            if detections is not None:
                emb_data = DeepFace.represent(
                    detections,
                    model_name="ArcFace",
                    detector_backend="retinaface",
                    enforce_detection=False
                )
                if emb_data:
                    emb = np.array(emb_data[0]["embedding"])
                    # Cosine similarity
                    similarities = [np.dot(emb, known_emb) / (np.linalg.norm(emb) * np.linalg.norm(known_emb))
                                    for known_emb in known_embeddings]
                    if similarities:
                        best_idx = np.argmax(similarities)
                        if similarities[best_idx] > 0.8:  # relaxed threshold
                            name = known_face_names[best_idx]
                            rollno = known_face_rollnos[best_idx]
                            mark_attendance(name, rollno)
                            cv2.putText(frame, f"{name} ({rollno})", (50, 50),
                                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                        else:
                            cv2.putText(frame, "Unknown", (50, 50),
                                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        except Exception as e:
            print(f"[ERROR] {e}")

        ret, buffer = cv2.imencode(".jpg", frame)
        frame = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')'''


def gen_frames():
    while True:
        success, frame = camera_cap.read()

        if not success:
            break

        try:
            # Detect faces using the recommended DeepFace function
            faces = DeepFace.extract_faces(
                img_path=frame,
                detector_backend="retinaface",
                enforce_detection=False,
                align=True
            )

            if faces:

                for face_data in faces:

                    face_img = face_data["face"]

                    # Convert float image to uint8
                    face_img = (face_img * 255).astype(np.uint8)

                    # Generate embedding for detected face
                    emb_data = DeepFace.represent(
                        img_path=face_img,
                        model_name="ArcFace",
                        detector_backend="skip",
                        enforce_detection=False
                    )

                    if emb_data:

                        emb = np.array(
                            emb_data[0]["embedding"]
                        )

                        similarities = []

                        for known_emb in known_embeddings:

                            similarity = np.dot(
                                emb,
                                known_emb
                            ) / (
                                np.linalg.norm(emb) *
                                np.linalg.norm(known_emb)
                            )

                            similarities.append(similarity)

                        if similarities:

                            best_idx = np.argmax(similarities)
                            best_similarity = similarities[best_idx]

                            print(
                                f"Best similarity: {best_similarity:.3f}"
                            )

                            if best_similarity > 0.40:

                                name = known_face_names[best_idx]
                                rollno = known_face_rollnos[best_idx]

                                mark_attendance(
                                    name,
                                    rollno
                                )

                                cv2.putText(
                                    frame,
                                    f"{name} ({rollno})",
                                    (50, 50),
                                    cv2.FONT_HERSHEY_SIMPLEX,
                                    1,
                                    (0, 255, 0),
                                    2
                                )

                            else:

                                cv2.putText(
                                    frame,
                                    "Unknown",
                                    (50, 50),
                                    cv2.FONT_HERSHEY_SIMPLEX,
                                    1,
                                    (0, 0, 255),
                                    2
                                )

        except Exception as e:

            print("[ERROR]", e)

        ret, buffer = cv2.imencode(
            ".jpg",
            frame
        )

        frame = buffer.tobytes()

        yield (
            b'--frame\r\n'
            b'Content-Type: image/jpeg\r\n\r\n'
            + frame +
            b'\r\n'
        )


# ============================
# Flask Routes
# ============================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video')
def video():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/admin')
def admin_page():
    cursor.execute("SELECT * FROM attendance")
    records = cursor.fetchall()
    return render_template('admin.html', records=records)

@app.route('/absent')
def absent_page():
    cursor.execute("SELECT rollno FROM attendance")
    present_rolls = [row[0] for row in cursor.fetchall()]
    absent_students = []
    for name, rollno in zip(known_face_names, known_face_rollnos):
        if rollno not in present_rolls:
            absent_students.append({"name": name, "rollno": rollno})
    return render_template('absent.html', absent_students=absent_students)

# ============================
# Run Flask App
# ============================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
